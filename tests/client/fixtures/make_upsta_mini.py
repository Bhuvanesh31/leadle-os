"""Generate tests/client/fixtures/upsta_mini.xlsx — tiny workbook for test_sheet_source_xlsx.

Run once: python tests/client/fixtures/make_upsta_mini.py
Commit the resulting upsta_mini.xlsx so tests never need to regenerate it.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

OUT = Path(__file__).parent / "upsta_mini.xlsx"


def _make_wb() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Spine tab: Prospect Data-US ──────────────────────────────────────────
    ws_spine = wb.create_sheet("Prospect Data-US")
    spine_header = [
        "Company Name", "Company Country", "Company Location",
        "Company Linked In URL", "Primary Industry", "Size (Text)",
        "Account Process", "First Name", "Last Name", "Full Name",
        "Location", "Company Domain", "LinkedIn Profile", "Title",
        "Email", "Phone", "Notes", "Start Date", "Status",
        "Aimfox ID", "Aimfox URN", "Instantly ID",
    ]
    ws_spine.append(spine_header)
    # Row 1: has both Aimfox ID and Instantly ID
    row1 = ["Acme Corp", "US", "San Francisco, CA", "https://linkedin.com/company/acme",
            "Software", "51-200", "Prospect", "John", "Doe", "John Doe",
            "San Francisco", "acme.com", "https://linkedin.com/in/johndoe", "CTO",
            "john@acme.com", "", "", "2026-01-01", "Active",
            229856678, "ACoAAA2zVaYBQXGrEKQf8TU4u1pSCTjzT45lh2Y",
            "019e89de-c4f3-7c29-9e9b-752159d50611"]
    ws_spine.append(row1)
    # Row 2: email-only (Instantly ID only, no Aimfox ID)
    row2 = ["Beta Inc", "US", "Austin, TX", "https://linkedin.com/company/beta",
            "Finance", "11-50", "Suspect", "Jane", "Smith", "Jane Smith",
            "Austin", "beta.com", "https://linkedin.com/in/janesmith", "CFO",
            "jane@beta.com", "", "", "2026-01-15", "Active",
            "", "", "019e89de-c4f3-7c29-9e9b-000000000001"]
    ws_spine.append(row2)

    # ── Singapore spine (second spine tab, just 1 row to keep it small) ─────
    ws_sg = wb.create_sheet("Prospect Data- Singapore")
    ws_sg.append(spine_header)
    row_sg = ["Gamma Ltd", "SG", "Singapore", "https://linkedin.com/company/gamma",
              "Logistics", "201-500", "High Priority Prospect",
              "Wei", "Tan", "Wei Tan", "Singapore", "gamma.sg",
              "https://linkedin.com/in/weitan", "CEO",
              "wei@gamma.sg", "", "", "", "Active",
              481204555, "ACoAABbbVaYBQXGrEKQf8ABCD", ""]
    ws_sg.append(row_sg)

    # ── Webhook - LinkedIn ───────────────────────────────────────────────────
    ws_li = wb.create_sheet("Webhook - LinkedIn")
    li_header = [
        "Event Type", "Company Name", "Profile Url", "Company Url",
        "Prospect Name", "Title", "Sender Profile", "Campaign Name",
        "Timestamp", "Date Extraction", "Connection request sent date",
        "Connection Accepted Date", "Reply Date", "Reply Messages",
        "Industry", "Size", "Company Country", "Process", "Variant",
        "Reply Sentiment",
    ]
    ws_li.append(li_header)
    # 1 reply row
    ts_li = datetime(2026, 6, 10, 9, 0, 0)
    reply_row = [
        "reply", "Acme Corp", "https://linkedin.com/in/johndoe",
        "https://linkedin.com/company/acme",
        "John Doe", "CTO", "https://linkedin.com/in/sender",
        "Upsta_US_PMP_V1", ts_li, "2026-06-10",
        "2026-05-01", "2026-05-10", "2026-06-10", "Sounds interesting",
        "Software", "51-200", "US", "Prospect", "", "neutral",
    ]
    ws_li.append(reply_row)
    # Repeated header row (pagination artifact — must be skipped)
    ws_li.append(li_header)

    # ── Webhook - Email ──────────────────────────────────────────────────────
    ws_em = wb.create_sheet("Webhook - Email")
    em_header = [
        "Company Name", "To Name", "Event Type", "Campaign Name",
        "Event Timestamp", "From Email", "To Email", "Reply Message",
        "Sent Date", "Email Open date", "Reply Date", "Email click date",
        "Sequence Number", "Title", "Website", "Company Size",
        "Company Country", "Industry", "Prospect Linkedin Profile",
        "Reply Sentiment",
    ]
    ws_em.append(em_header)
    ts_open1 = datetime(2026, 6, 12, 10, 30, 0)
    ts_open2 = datetime(2026, 6, 13, 14, 0, 0)
    ts_sent = datetime(2026, 6, 11, 8, 0, 0)
    ws_em.append([
        "Acme Corp", "John Doe", "email_opened", "Upsta_SFDI_V1",
        ts_open1, "sender@leadle.in", "john@acme.com", "",
        "2026-06-11", "2026-06-12", "", "", "1", "CTO",
        "acme.com", "51-200", "US", "Software",
        "https://linkedin.com/in/johndoe", "",
    ])
    ws_em.append([
        "Beta Inc", "Jane Smith", "email_opened", "Upsta_SFDI_V1",
        ts_open2, "sender@leadle.in", "jane@beta.com", "",
        "2026-06-11", "2026-06-13", "", "", "1", "CFO",
        "beta.com", "11-50", "US", "Finance",
        "https://linkedin.com/in/janesmith", "",
    ])
    # email_sent row — should NOT appear in opens
    ws_em.append([
        "Beta Inc", "Jane Smith", "email_sent", "Upsta_SFDI_V1",
        ts_sent, "sender@leadle.in", "jane@beta.com", "",
        "2026-06-11", "", "", "", "1", "CFO",
        "beta.com", "11-50", "US", "Finance",
        "https://linkedin.com/in/janesmith", "",
    ])

    # ── Response Tracker ─────────────────────────────────────────────────────
    ws_resp = wb.create_sheet("Response Tracker")
    resp_header = [
        "Channel", "Account", "Response Date", "Status", "Response",
        "LinkedIn", "Name", "Job Title", "Company", "Company Url",
        "Company Web", "Loc",
    ]
    ws_resp.append(resp_header)
    ws_resp.append([
        "LinkedIn", "Acme Corp", "2026-06-10", "Interested",
        "Yes let's connect next week",
        "https://linkedin.com/in/johndoe", "John Doe", "CTO",
        "Acme Corp", "https://linkedin.com/company/acme",
        "acme.com", "San Francisco",
    ])

    return wb


if __name__ == "__main__":
    wb = _make_wb()
    wb.save(OUT)
    print(f"Written: {OUT}")
