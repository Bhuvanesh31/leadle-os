"""read_sheets must produce the same ClientData as read_xlsx for the same rows."""

from pathlib import Path

import openpyxl

from dashboard.client.sources import sheet_source

FIX = Path(__file__).parent / "fixtures" / "upsta_mini.xlsx"


class _FakeWorksheet:
    def __init__(self, title, rows):
        self.title, self._rows = title, rows

    def get_all_values(self):
        return self._rows


class _FakeSpreadsheet:
    def __init__(self, sheets):
        self._sheets = sheets

    def worksheets(self):
        return [_FakeWorksheet(t, r) for t, r in self._sheets.items()]

    def worksheet(self, title):
        return _FakeWorksheet(title, self._sheets[title])


class _FakeClient:
    def __init__(self, sheets):
        self._sheets = sheets

    def open_by_key(self, key):
        return _FakeSpreadsheet(self._sheets)


def _fixture_tabs():
    wb = openpyxl.load_workbook(str(FIX), read_only=True, data_only=True)
    tabs = {}
    for name in sheet_source._ALL_SHEET_TABS:
        if name in wb.sheetnames:
            tabs[name] = [list(r) for r in wb[name].iter_rows(values_only=True)]
    wb.close()
    return tabs


def test_read_sheets_matches_read_xlsx():
    fake = _FakeClient(_fixture_tabs())
    from_sheets = sheet_source.read_sheets("any-id", client=fake)
    from_xlsx = sheet_source.read_xlsx(str(FIX))

    # Precondition: the fixture must carry real data, else the parity asserts
    # below pass vacuously (0 == 0, {} == {}) and prove nothing.
    assert from_xlsx.targets and from_xlsx.replies and from_xlsx.opens and from_xlsx.warm_leads

    assert len(from_sheets.targets) == len(from_xlsx.targets)
    assert len(from_sheets.replies) == len(from_xlsx.replies)
    assert len(from_sheets.opens) == len(from_xlsx.opens)
    assert len(from_sheets.warm_leads) == len(from_xlsx.warm_leads)
    assert {t.aimfox_id for t in from_sheets.targets} == {t.aimfox_id for t in from_xlsx.targets}
    assert {r.sentiment for r in from_sheets.replies} == {r.sentiment for r in from_xlsx.replies}
