"""v1 source: parse the Drive-dumped 'Prospect list' workbook text into ClientData.

The workbook arrives as the markdown-table text emitted by the Drive MCP
read_file_content tool (the session dumps it to a file; this module reads that
file). Tables are delimited by their header row; underscores may be backslash-
escaped in the dump, so we strip backslashes from every cell.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from dashboard.client.model import (
    ClientData,
    Context,
    EmailEvent,
    LinkedInEvent,
    TargetCo,
    WarmLead,
)

# Header signatures (first few columns) that mark the start of each table.
_H_RESP = "| Channel | Account | Response Date | Status"
_H_TARGET = "| Company Name | Company Country | Company Location"
_H_LI = "| Event Type | Company Name | Profile Url"
_H_EMAIL = "| Company Name | To Name | Event Type | Campaign Name"
_H_ICP = "| Column 1 | Offering to the market"
_ALL_HEADERS = (_H_RESP, _H_TARGET, _H_LI, _H_EMAIL, _H_ICP,
                "| Item | Status | Responsibility")


def _cells(line: str) -> list[str]:
    return [c.strip().replace("\\", "") for c in line.strip().strip("|").split("|")]


def _is_sep(line: str) -> bool:
    return set(line.replace("|", "").replace(" ", "")) <= set(":-")


def _tables_under(lines: list[str], header_prefix: str):
    """Yield (header_cells, data_rows) for EVERY table whose header starts with
    header_prefix. The Drive flatten re-emits the header for each paginated block."""
    i, n = 0, len(lines)
    while i < n:
        if lines[i].startswith(header_prefix):
            header = _cells(lines[i])
            i += 1
            rows: list[list[str]] = []
            while i < n:
                ln = lines[i]
                if not ln.strip().startswith("|"):
                    i += 1
                    continue
                if _is_sep(ln):
                    i += 1
                    continue
                if any(ln.startswith(h) for h in _ALL_HEADERS):
                    break  # next table (possibly the same header = next page)
                rows.append(_cells(ln))
                i += 1
            yield header, rows
        else:
            i += 1


def _rows_under(lines: list[str], header_prefix: str) -> list[list[str]]:
    """Data rows across ALL tables under header_prefix, concatenated."""
    out: list[list[str]] = []
    for _header, rows in _tables_under(lines, header_prefix):
        out.extend(rows)
    return out


def _col(header: list[str], *names: str) -> int:
    for nm in names:
        if nm in header:
            return header.index(nm)
    return -1


def _g(row: list[str], idx: int) -> str:
    return row[idx] if 0 <= idx < len(row) else ""


def parse(workbook_text: str, client: str) -> ClientData:
    lines = workbook_text.split("\n")
    pfx = f"{client.lower()}_"

    emails: list[EmailEvent] = []
    for r in _rows_under(lines, _H_EMAIL):
        if len(r) < 6 or r[2] in ("", "Event Type"):
            continue
        campaign = r[3]
        if not campaign.lower().startswith(pfx):
            continue
        try:
            ts = datetime.fromisoformat(r[4].replace("Z", "+00:00"))
        except ValueError:
            continue
        emails.append(EmailEvent(r[0], r[1], r[2], campaign, ts, r[5]))

    linkedin: list[LinkedInEvent] = []
    for r in _rows_under(lines, _H_LI):
        if len(r) < 6 or r[0] in ("", "Event Type"):
            continue
        linkedin.append(LinkedInEvent(r[0], r[1], r[2], r[4], r[5]))

    warm: list[WarmLead] = []
    for r in _rows_under(lines, _H_RESP):
        if len(r) < 12 or r[0] in ("", "Channel"):
            continue
        # Columns: Channel(0) Account(1) ResponseDate(2) Status(3) Response(4)
        # LinkedIn(5) Name(6) JobTitle(7) Company(8) CompanyUrl(9) CompanyWeb(10) Loc(11)
        # WarmLead has 11 fields; skip CompanyWeb (index 10), use Loc (index 11) as location.
        warm.append(WarmLead(*r[:10], r[11]))

    targets: list[TargetCo] = []
    for header, rows in _tables_under(lines, _H_TARGET):
        c_name = _col(header, "Company Name")
        c_country = _col(header, "Company Country")
        c_loc = _col(header, "Company Location")
        c_li = _col(header, "Company Linked In URL", "Company LinkedIn URL")
        c_ind = _col(header, "Primary Industry")
        c_size = _col(header, "Size (Text)", "Size")
        c_seg = _col(header, "Account Process")
        c_dom = _col(header, "Company Domain")
        c_af = _col(header, "Aimfox ID")
        c_urn = _col(header, "Aimfox URN")
        c_inst = _col(header, "Instantly ID")
        for r in rows:
            name = _g(r, c_name)
            if name in ("", "Company Name"):
                continue
            targets.append(TargetCo(
                name, _g(r, c_country), _g(r, c_loc), _g(r, c_li), _g(r, c_ind),
                _g(r, c_size), _g(r, c_seg), _g(r, c_dom),
                aimfox_id=_g(r, c_af), aimfox_urn=_g(r, c_urn),
                instantly_id=_g(r, c_inst)))

    channels: list[str] = []
    for r in _rows_under(lines, _H_ICP):
        if len(r) >= 3 and r[2]:
            channels = [c.strip() for c in r[2].split(",") if c.strip()]
            break

    ctx = Context(client=client, channels=channels, campaign_live_dates={}, icp={})
    return ClientData(emails, linkedin, warm, targets, ctx)


def read(client: str, workbook_path: str) -> ClientData:
    return parse(Path(workbook_path).read_text(encoding="utf-8"), client)


# ─────────────────────────────────────────────────────────────────────────────
# XLSX ingestion (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

_SPINE_TABS = ["Prospect Data-US", "Prospect Data- Singapore"]
_WH_LINKEDIN = "Webhook - LinkedIn"
_WH_EMAIL = "Webhook - Email"
_RESP_TAB = "Response Tracker"

# LinkedIn reply Event Types that indicate a human reply
_LI_REPLY_TYPES = {"reply", "campaign_reply"}


def _get(row_vals: list, col_map: dict[str, int], *names: str) -> str:
    """Return first non-None value for any of *names; empty string when absent."""
    for name in names:
        idx = col_map.get(name)
        if idx is not None and idx < len(row_vals):
            v = row_vals[idx]
            if v is not None:
                return str(v).strip()
    return ""


def _coerce_aimfox_id(raw: str) -> str:
    """openpyxl reads numeric cells as float; '229856678.0' -> '229856678'."""
    if not raw:
        return raw
    try:
        return str(int(float(raw)))
    except (ValueError, OverflowError):
        return raw


def _parse_ts(v) -> datetime | None:
    """Coerce openpyxl cell value to datetime. Handles datetime objects and ISO strings."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S+00:00",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _is_header_repeat(row: tuple, header_tuple: tuple) -> bool:
    """Return True if row is a repeated emission of the header (pagination artifact).

    Compares normalised string values of both tuples so the check is exact.
    """
    row_strs = tuple(str(v).strip() if v is not None else "" for v in row)
    hdr_strs = tuple(str(v).strip() if v is not None else "" for v in header_tuple)
    return row_strs == hdr_strs


def read_xlsx(path: str) -> ClientData:
    """Parse the client prospect XLSX workbook and return a ClientData.

    Uses openpyxl(read_only=True, data_only=True). Resolves columns by header
    NAME (not position). Skips repeated header rows (pagination artifact) and
    all-None rows. Int-coerces numeric Aimfox IDs to avoid float artifacts.
    """
    import openpyxl  # local import — not always in the text-parse path

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    targets: list[TargetCo] = []
    replies: list = []
    opens: list = []
    warm_leads: list[WarmLead] = []

    # ── Spine tabs ────────────────────────────────────────────────────────────
    for tab_name in _SPINE_TABS:
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            continue
        col_map = {
            str(v).strip(): idx
            for idx, v in enumerate(header_row)
            if v is not None
        }
        for row in rows:
            if all(v is None for v in row):
                continue
            if _is_header_repeat(row, header_row):
                continue
            rv = list(row)
            name = _get(rv, col_map, "Company Name")
            if not name:
                continue
            raw_af = _get(rv, col_map, "Aimfox ID")
            targets.append(TargetCo(
                name=name,
                country=_get(rv, col_map, "Company Country"),
                location=_get(rv, col_map, "Company Location"),
                linkedin_url=_get(rv, col_map, "Company Linked In URL", "Company LinkedIn URL"),
                industry=_get(rv, col_map, "Primary Industry"),
                size=_get(rv, col_map, "Size (Text)", "Size"),
                segment=_get(rv, col_map, "Account Process"),
                domain=_get(rv, col_map, "Company Domain"),
                aimfox_id=_coerce_aimfox_id(raw_af),
                aimfox_urn=_get(rv, col_map, "Aimfox URN"),
                instantly_id=_get(rv, col_map, "Instantly ID"),
            ))

    # ── Webhook - LinkedIn ────────────────────────────────────────────────────
    if _WH_LINKEDIN in wb.sheetnames:
        ws = wb[_WH_LINKEDIN]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is not None:
            col_map = {
                str(v).strip(): idx
                for idx, v in enumerate(header_row)
                if v is not None
            }
            from dashboard.client.model import ReplyRecord
            for row in rows_iter:
                if all(v is None for v in row):
                    continue
                if _is_header_repeat(row, header_row):
                    continue
                rv = list(row)
                evt = _get(rv, col_map, "Event Type").lower()
                if evt not in _LI_REPLY_TYPES:
                    continue
                sentiment = _get(rv, col_map, "Reply Sentiment") or "untagged"
                campaign = _get(rv, col_map, "Campaign Name")
                name = _get(rv, col_map, "Prospect Name")
                raw_ts = _get(rv, col_map, "Timestamp") if "Timestamp" in col_map else None
                ts = _parse_ts(raw_ts)
                replies.append(ReplyRecord(
                    channel="linkedin",
                    campaign=campaign,
                    sentiment=sentiment,
                    name=name,
                    ts=ts,
                ))

    # ── Webhook - Email ───────────────────────────────────────────────────────
    if _WH_EMAIL in wb.sheetnames:
        ws = wb[_WH_EMAIL]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is not None:
            col_map = {
                str(v).strip(): idx
                for idx, v in enumerate(header_row)
                if v is not None
            }
            from dashboard.client.model import OpenEvent
            for row in rows_iter:
                if all(v is None for v in row):
                    continue
                if _is_header_repeat(row, header_row):
                    continue
                rv = list(row)
                evt = _get(rv, col_map, "Event Type").lower()
                if evt == "email_opened":
                    raw_ts = _get(rv, col_map, "Event Timestamp") if "Event Timestamp" in col_map else None
                    ts = _parse_ts(raw_ts)
                    if ts is not None:
                        opens.append(OpenEvent(channel="email", ts=ts))

    # ── Response Tracker ─────────────────────────────────────────────────────
    if _RESP_TAB in wb.sheetnames:
        ws = wb[_RESP_TAB]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is not None:
            col_map = {
                str(v).strip(): idx
                for idx, v in enumerate(header_row)
                if v is not None
            }
            for row in rows_iter:
                if all(v is None for v in row):
                    continue
                if _is_header_repeat(row, header_row):
                    continue
                rv = list(row)
                channel = _get(rv, col_map, "Channel")
                if not channel:
                    continue
                warm_leads.append(WarmLead(
                    channel=channel,
                    account=_get(rv, col_map, "Account"),
                    response_date=_get(rv, col_map, "Response Date"),
                    status=_get(rv, col_map, "Status"),
                    response_text=_get(rv, col_map, "Response"),
                    linkedin_url=_get(rv, col_map, "LinkedIn"),
                    name=_get(rv, col_map, "Name"),
                    title=_get(rv, col_map, "Job Title"),
                    company=_get(rv, col_map, "Company"),
                    company_url=_get(rv, col_map, "Company Url"),
                    location=_get(rv, col_map, "Loc"),
                ))

    wb.close()
    return ClientData(targets=targets, replies=replies, opens=opens, warm_leads=warm_leads)
